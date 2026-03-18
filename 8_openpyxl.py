

# مشروع from openpyxl import load_workbook
from openpyxl import load_workbook
file_path = "C:\\..\\studunts.xlsx" # تحديد مسار الملف

workbook = load_workbook(file_path)  # فتح الملف

# اختيار ورقة العمل (Sheet)

sheet = workbook.active
sheeta = workbook["Sheet1"]
sheeta["A1"]="q"



# إضافة بيانات باستخدام حلقات
data = [("Ali", 10), ("Sara", 12), ("Laila", 5)]

for row_num, row_data in enumerate(data, start=6):
    for col_num, value in enumerate(row_data, start=1):
        sheet.cell(row=row_num, column=col_num).value = value
# sheet["A2"]= None
sheet.delete_rows(2)

# قراءة البيانات
for row in sheet.iter_rows(values_only=True):
    print(row)
for row in sheeta.iter_rows(values_only=True):
    print(row)
print('----------------------')

# اختيار النطاق A1:B2
for row in sheet["A3:B4"]:
    for cell in row:
        print(cell.value)
print('----------------------')
# المرور على كل الصفوف والخلايا في نطاق معين
# min_row: أول صف تبدأ منه.
# max_row: آخر صف تتوقف عنده.
# min_col: أول عمود تبدأ منه.
# max_col: آخر عمود تتوقف عنده.
for row in sheet.iter_rows(min_row=1, max_row=5, min_col=1, max_col=3):
    for cell in row:
        print(cell.value)
print('----------------------')
print(sheet["A1"].value)



workbook.save(file_path)




# مشروع import openpyxl
import openpyxl
from openpyxl import load_workbook

t= open("C:\\..\\nowsum.xlsx")
# print(bin(ord('6')))
# print(bin(6))
x= open(r"C:\..\nowsum.txt",'a')
y= open("C:/../nowsum.txt",'w') # if not founded python will founded
print(x.readable())  # are the file rdy to read
# print(x.read())      # read all file
# print(x.readline())  # read line thin line if tow print will print line1 thin line2
# print(x.readlines()) # print all lines in list and can add [2] to read line3
# for i in x.readlines():
#     print(i)
# num= [3,2,5,6]
# y.write("\njgjhvgh") # add
# y.writelines(num)    # add list num

x.close()
y.close()



print('____________________________________________ مشروع')

# مشروع from openpyxl import load_workbook
# اكسل لتعبأة اسماء الطلاب
# يطلب من المستخدم ادخال ثلاث درجات لكل طالب
# احسب متوسط دجات كل طالب
# H4 طباعة اسم الطالب الاعلى نسبة عند مناداة 

from openpyxl import load_workbook
file_path =	  r'C:/../studunts.xlsx'
workbook  = load_workbook(file_path)
sheet = workbook["Studunt"]

# file_patha =  r'C:/../newsulaiman.xlsx'  # test
# workbooka  = load_workbook(file_patha)  # test

def students():
    for i in range(1,4):  # هذه لاظافة الاسماء
        A= input(f"enter studant {i} name: ")
        sheet[f"A{i}"] = A

    studant_ave=0
    xl_row= ['B','C','D']
    for i in range(1,4):   # هذه اختيار اول طالب
        for row in xl_row: # هنا لتعبأة ثلاث درجات للطالب
            studantName= sheet[f"A{i}"].value
            b = int(input(f"enter grades studant {studantName}: "))
            sheet[f"{row}{i}"]=b
            studant_ave+= sheet[f"{row}{i}"].value
        studant_ave/=3
        sheet[f"E{i}"] = studant_ave

    avNum,iNum = 0,0
    for i in sheet["E1:E3"]: # هنا يحفظ اكبر متوسط من الدرجات في الخلية H4
        for c in i:
            iNum+=1
            if c.value > avNum:
                avNum = c.value
                avNam = sheet[f"A{iNum}"].value
    sheet["H5"]= avNum
    sheet["H4"] = avNam

students()
print(f"name top of the class is {sheet["H4"].value} with average grades {sheet["H5"].value} ")
workbook.save(file_path)



